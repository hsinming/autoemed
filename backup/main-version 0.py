#!/usr/bin/env python3
# -*- coding:utf-8 -*-

# nuitka-project: --mingw64
# nuitka-project: --show-progress
# nuitka-project: --show-memory
# nuitka-project: --standalone
# nuitka-project: --windows-console-mode=disable
# nuitka-project: --windows-icon-from-ico=icon.ico
# nuitka-project: --output-dir=build

# 启用插件
# nuitka-project: --enable-plugin=tk-inter
# nuitka-project: --enable-plugin=upx

# 移除不必要的导入
# nuitka-project: --nofollow-import-to=numpy
# nuitka-project: --nofollow-import-to=pandas
"""
@author: Hsin-ming Chen
@license: MIT
@file: main.py
@time: 2025/02/06
@contact: hsinming.chen@gmail.com
@software: PyCharm
"""
import logging
from pathlib import Path
from time import sleep
import threading
from openpyxl import load_workbook
from helium import start_chrome, write, click, wait_until, find_all, kill_browser, Text, TextField, Button, RadioButton, CheckBox, Alert
from selenium.webdriver import ChromeOptions
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
import tkinter as tk
from tkinter import ttk, filedialog, StringVar, BooleanVar

VERSION = "1.5"
EMEDICAL_URL = 'https://www.emedical.immi.gov.au/eMedUI/eMedical'
MAX_LOGIN_ATTEMPTS = 1

# 設定日誌
log_file = Path("log.txt")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler(log_file, mode='a', encoding='utf-8'),
    logging.StreamHandler()
])

# 中止標誌 (使用 threading.Event)
stop_event = threading.Event()


def extract_emedical_no(file_path):
    """從 Excel 讀取 eMedical No. 清單，並過濾出黑色文字且無底色的數據"""
    def is_black_font(cell):
        """檢查單元格文字是否為黑色"""
        return (
            cell.font is None or
            cell.font.color is None or
            cell.font.color.rgb in ["FF000000", None]  # None 表示未設置顏色，預設為黑色
        )

    def is_no_fill(cell):
        """檢查單元格是否沒有填充背景色"""
        return (
            cell.fill is None or
            cell.fill.fgColor is None or
            cell.fill.fgColor.rgb in ["00000000", "FFFFFFFF", None]  # 透明 or 白色
        )

    if not Path(file_path).exists():
        logging.error(f"找不到 Excel 檔案: {file_path}")
        return []

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        logging.error(f"讀取 Excel 發生錯誤: {e}")
        return []

    all_emedical_nos = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        found_header = False  # 確保有找到 "eMedical No."
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip() == "eMedical No.":
                    found_header = True
                    # 獲取此列下方的所有值
                    col_idx = cell.column
                    for r in range(cell.row + 1, ws.max_row + 1):
                        target_cell = ws.cell(row=r, column=col_idx)
                        if target_cell.value and is_black_font(target_cell) and is_no_fill(target_cell):
                            all_emedical_nos.append(str(target_cell.value))

        if not found_header:
            logging.warning(f"無法在 {sheet_name} 找到 'eMedical No.' 欄位")

    wb.close()

    if len(all_emedical_nos) == 0:
        logging.warning("未讀取到任何有效的 eMedical No.")

    return all_emedical_nos


def login_to_emedical(user_id, password, headless):
    options = ChromeOptions()
    options.add_argument("--disable-extensions")  # 禁用擴展，防止擴展產生 scoped_dir
    options.add_argument("--incognito")  # 無痕模式，減少緩存
    options.add_argument("--no-sandbox")  # 避免某些安全性限制
    options.add_argument("--disable-gpu")  # 減少 WebGPU 緩存
    options.add_argument("--disable-background-networking")  # 停止後台更新
    options.add_argument("--disable-component-update")  # 禁用組件更新
    options.add_argument("--disable-features=NetworkService,NetworkServiceInProcess")  # 減少網絡相關的暫存文件

    for attempt in range(1, MAX_LOGIN_ATTEMPTS + 1):
        try:
            logging.info("啟動瀏覽器並登入 eMedical 系統")
            start_chrome(EMEDICAL_URL, headless=headless, options=options)
            write(user_id, into=TextField('User id'))
            write(password, into=TextField('Password'))
            click(Button('Logon'))
            wait_until(Text('Case search').exists, timeout_secs=10)
            logging.info("登入成功！")
            return True  # 登入成功

        except NoSuchElementException:
            logging.error("找不到登入欄位，請檢查頁面是否變更")

        except TimeoutException:
            logging.warning(f"登入超時，嘗試第 {attempt} 次...")
            sleep(2)  # 等待後重試

        except WebDriverException as e:
            logging.error(f"瀏覽器發生錯誤: {e}")
            break

    logging.error("多次嘗試登入失敗，請檢查帳號密碼")
    return False


def emedical_cxr_automation(emed_no: str, country: str) -> bool:
    """
    通用的 eMedical 502 Chest X-Ray正常案例自動點選流程處理函數
    根據 country 參數決定是否有額外的步驟
    """
    try:
        if not RadioButton('Using Health Case Identifier').is_selected():
            click(RadioButton('Using Health Case Identifier'))

        write(emed_no, into=TextField('ID'))
        click(Button('Search', to_right_of=Button('Reset')))
        wait_until(Text('Select:').exists)
        sleep(1)    #wait for reading status
        click(Button('All'))
        click(Button('Manage Case'))
        wait_until(Text('Pre exam: Health case details').exists)

        if Text('502 Chest X-Ray Examination').exists():
            click(Text('502 Chest X-Ray Examination'))

            # 依國家需求點擊不同的按鈕
            if country == "美國":
                click(Text('Findings'))
                wait_until(Text('502 Chest X-Ray Examination: Findings').exists)

                if not RadioButton('Normal', to_right_of=Text('Findings')).is_selected():
                    click(RadioButton('Normal', to_right_of=Text('Findings')))
            else:
                click(Text('Detailed radiology findings'))
                wait_until(Text('Detailed question').exists)

                for normal_button in find_all(RadioButton('Normal')):
                    if not normal_button.is_selected():
                        click(normal_button)

                if not RadioButton('Absent').is_selected():
                    click(RadioButton('Absent'))

                if not RadioButton('No',
                                   to_right_of=RadioButton('Not selected', to_right_of=Text('7. Are there strong suspicions of active Tuberculosis (TB)?'))).is_selected():
                    click(RadioButton('No',
                                      to_right_of=RadioButton('Not selected', to_right_of=Text('7. Are there strong suspicions of active Tuberculosis (TB)?'))))

                if country == "加拿大":
                    click(Button('Next'))
                    wait_until(Text('Special findings').exists)

                    if not RadioButton('None of the following are present').is_selected():
                        click(RadioButton('None of the following are present'))

            click(Button('Next'))
            wait_until(Text('502 Chest X-Ray Examination: Review exam details').exists)
            sleep(1)
            click(Button('Next'))

            if country == "美國":
                wait_until(Text('502 Chest X-Ray Examination: Examiner Declaration').exists)
                if Button('Prepare for declaration').exists() and Button('Prepare for declaration').is_enabled():
                    click(Button('Prepare for declaration'))
            else:
                wait_until(Text('502 Chest X-Ray Examination: Grading & Examiner Declaration').exists)
                if Button('Prepare for grading').exists() and Button('Prepare for grading').is_enabled():
                    click(Button('Prepare for grading'))

            wait_until(Text('Examiner declaration').exists)
            if not CheckBox(
                    'I declare that the chest X-ray examination report is a true and correct record of my findings.').is_checked():
                click(CheckBox(
                    'I declare that the chest X-ray examination report is a true and correct record of my findings.'))

            if country != "美國":
                if not RadioButton(
                        'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.').is_selected():
                    click(RadioButton(
                        'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.'))

            if Button('Submit Exam').exists() and Button('Submit Exam').is_enabled():
                click(Button('Submit Exam'))
                wait_until(Alert().exists)
                Alert().accept()
                wait_until(Text('Success').exists)

        click(Button('Close'))
        logging.info(f"成功處理 ({country}): {emed_no}")
        return True

    except (NoSuchElementException, TimeoutException, WebDriverException) as e:
        logging.error(f"自動化失敗: {emed_no}, 錯誤: {e}")
        # 嘗試關閉目前的視窗，避免卡住
        try:
            if Button('Close').exists() and Button('Close').is_enabled():
                click(Button('Close'))
        except Exception:
            pass
        return False


def get_country(emed_no):
    """根據 eMedical No. 的前綴判斷對應的國家"""
    prefix_to_country = {
        "HAP": "澳大利亞",
        "TRN": "澳大利亞",
        "NZER": "紐西蘭",
        "NZHR": "紐西蘭",
        "IME": "加拿大",
        "UMI": "加拿大",
        "UCI": "加拿大",
        "CEAC": "美國",
    }

    for prefix, country in prefix_to_country.items():
        if emed_no.startswith(prefix):
            return country

    return "未知國家"


def emedical_workflow(user_id, password, excel_path, update_status, emed_no_listbox, success_listbox,
                      failure_listbox, headless, close_browser, update_counts):
    logging.info(f"讀取 eMedical No. 來自 {excel_path}")
    if not Path(excel_path).exists():
        logging.error(f"找不到檔案: {excel_path}")
        update_status(f"錯誤: 找不到檔案 {excel_path}")
        return  # 直接結束函式

    emedical_numbers = extract_emedical_no(Path(excel_path))
    if not emedical_numbers:
        logging.warning("未讀取到任何 eMedical No.")
        update_status("未讀取到任何 eMedical No.")
        return  # 直接結束函式
    logging.info(f"讀取 {len(emedical_numbers)} 個 eMedical No.")

    # 將 eMedical No. 加入 GUI 清單
    for emed_no in emedical_numbers:
        emed_no_listbox.insert(tk.END, emed_no)

    if not login_to_emedical(user_id, password, headless):
        update_status("登入失敗，請檢查帳號密碼")
        return  # 終止執行

    for index, emed_no in enumerate(emedical_numbers):
        if stop_event.is_set():
            logging.info("用戶手動中止處理")
            update_status("處理已中止")
            break

        update_status(f'現在處理: {emed_no}')
        logging.info(f'現在處理: {emed_no}')

        # 標記當前處理中的 eMedical No.
        emed_no_listbox.itemconfig(index, {'bg': 'blue', 'fg': 'white'})

        country = get_country(emed_no)
        success = False  # 預設為失敗

        if country == "未知國家":
            logging.warning(f"未知國家的 eMedical No.: {emed_no}")
        else:
            success = emedical_cxr_automation(emed_no, country)

        emed_no_listbox.itemconfig(index, {'bg': 'white', 'fg': 'black'})  # 恢復顏色

        if success:
            success_listbox.insert(tk.END, emed_no)
        else:
            failure_listbox.insert(tk.END, emed_no)

        update_counts()

    update_status(f"處理完成！成功: {success_listbox.size()}, 失敗: {failure_listbox.size()}")
    logging.info(f"處理完成！成功: {success_listbox.size()}, 失敗: {failure_listbox.size()}")

    if close_browser or headless:
        logging.info("關閉瀏覽器")
        kill_browser()


def start_gui():
    root = tk.Tk()
    root.title(f"eMedical Automation v{VERSION} By HsinMing Chen")
    root.geometry("500x700")  # 預設視窗大小
    root.minsize(500, 700)  # 最小大小，避免視窗過小
    root.resizable(True, True)  # 允許調整大小

    style = ttk.Style()
    style.configure("TButton", font=("Arial", 12))
    style.configure("TLabel", font=("Arial", 11))
    style.configure("TEntry", font=("Arial", 11))

    main_frame = ttk.Frame(root, padding=10)
    main_frame.grid(row=0, column=0, sticky="nsew")

    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    # 設置 Grid 可變動大小
    for i in range(8):
        main_frame.rowconfigure(i, weight=1)
    main_frame.columnconfigure(0, weight=1)

    # 使用者資訊區塊
    user_frame = ttk.LabelFrame(main_frame, text="User Login", padding=10)
    user_frame.grid(row=0, column=0, sticky="nsew", pady=5)

    ttk.Label(user_frame, text="User ID:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    user_id_var = StringVar()
    ttk.Entry(user_frame, textvariable=user_id_var, width=30).grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(user_frame, text="Password:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    password_var = StringVar()
    ttk.Entry(user_frame, textvariable=password_var, show='*', width=30).grid(row=1, column=1, padx=5, pady=5)

    # Excel 檔案選擇區塊
    file_frame = ttk.LabelFrame(main_frame, text="Excel File", padding=10)
    file_frame.grid(row=1, column=0, sticky="nsew", pady=5)

    excel_path_var = StringVar()
    file_entry = ttk.Entry(file_frame, textvariable=excel_path_var)
    file_entry.grid(row=0, column=0, sticky="ew", padx=5)

    def select_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        excel_path_var.set(file_path)

    ttk.Button(file_frame, text="Browse", command=select_file).grid(row=0, column=1, padx=5, sticky="e")

    file_frame.columnconfigure(0, weight=1)  # 讓輸入框可以隨視窗變動

    # 選項區塊
    options_frame = ttk.LabelFrame(main_frame, text="Options", padding=10)
    options_frame.grid(row=2, column=0, sticky="nsew", pady=5)

    headless_var = BooleanVar()
    close_browser_var = BooleanVar()
    ttk.Checkbutton(options_frame, text="Headless Mode", variable=headless_var).grid(row=0, column=0, sticky="w")
    ttk.Checkbutton(options_frame, text="Kill Browser After Completion", variable=close_browser_var).grid(row=1,
                                                                                                          column=0,
                                                                                                          sticky="w")

    # 狀態顯示
    status_var = StringVar()
    ttk.Label(main_frame, textvariable=status_var, foreground='blue').grid(row=3, column=0, pady=5, sticky="w")

    def update_status(msg):
        status_var.set(msg)
        root.after(100, lambda: root.update_idletasks())  # 避免 GUI 卡住

    # eMedical No. Listbox
    list_frame = ttk.LabelFrame(main_frame, text="eMedical No. 清單", padding=10)
    list_frame.grid(row=4, column=0, sticky="nsew", pady=5)

    emed_no_listbox = tk.Listbox(list_frame, height=5)
    emed_no_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=emed_no_listbox.yview)
    emed_no_listbox.config(yscrollcommand=emed_no_scrollbar.set)
    emed_no_listbox.grid(row=0, column=0, sticky="nsew")
    emed_no_scrollbar.grid(row=0, column=1, sticky="ns")

    list_frame.rowconfigure(0, weight=1)
    list_frame.columnconfigure(0, weight=1)

    # 成功 & 失敗 ListBox
    result_frame = ttk.LabelFrame(main_frame, text="Processing Results", padding=10)
    result_frame.grid(row=5, column=0, sticky="nsew", pady=5)

    ttk.Label(result_frame, text="成功的 eMedical No. (數量: 0)", name="success_label").grid(row=0, column=0,
                                                                                             sticky="w")
    success_listbox = tk.Listbox(result_frame, height=5)
    success_listbox.grid(row=1, column=0, sticky="nsew", padx=5, pady=2)

    ttk.Label(result_frame, text="失敗的 eMedical No. (數量: 0)", name="failure_label").grid(row=2, column=0,
                                                                                             sticky="w")
    failure_listbox = tk.Listbox(result_frame, height=5)
    failure_listbox.grid(row=3, column=0, sticky="nsew", padx=5, pady=2)

    result_frame.rowconfigure(1, weight=1)
    result_frame.rowconfigure(3, weight=1)
    result_frame.columnconfigure(0, weight=1)

    def update_counts():
        success_count = success_listbox.size()
        failure_count = failure_listbox.size()
        success_label = result_frame.nametowidget("success_label")
        failure_label = result_frame.nametowidget("failure_label")
        success_label.config(text=f"成功的 eMedical No. (數量: {success_count})")
        failure_label.config(text=f"失敗的 eMedical No. (數量: {failure_count})")

    def start_emedical_workflow():
        user_id = user_id_var.get()
        password = password_var.get()
        excel_path = excel_path_var.get()
        headless = headless_var.get()
        close_browser = close_browser_var.get()

        if not user_id or not password or not excel_path:
            update_status("請填寫所有欄位！")
            return

        # 清空 ListBox 內容
        emed_no_listbox.delete(0, tk.END)
        success_listbox.delete(0, tk.END)
        failure_listbox.delete(0, tk.END)
        update_counts()

        update_status("正在處理...")
        stop_event.clear()

        # **使用 Thread 避免 GUI 被阻塞**
        emedical_workflow_thread = threading.Thread(
            target=emedical_workflow,
            args=(user_id, password, excel_path, update_status, emed_no_listbox, success_listbox,
                  failure_listbox, headless, close_browser, update_counts),
            daemon=True  # 設為 Daemon，確保主程式關閉時該 Thread 也會終止
        )
        emedical_workflow_thread.start()

    def stop_emedical_workflow():
        """中止處理流程"""
        stop_event.set()
        update_status("已中止處理")

    # 按鈕區塊
    button_frame = ttk.Frame(main_frame, padding=10)
    button_frame.grid(row=6, column=0, sticky="nsew", pady=5)

    ttk.Button(button_frame, text="Start", command=start_emedical_workflow).grid(row=0, column=0, sticky="ew")
    ttk.Button(button_frame, text="Stop", command=stop_emedical_workflow).grid(row=0, column=1, sticky="ew")

    button_frame.columnconfigure(0, weight=1)
    button_frame.columnconfigure(1, weight=1)

    root.mainloop()


if __name__ == "__main__":
    start_gui()

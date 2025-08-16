#!/usr/bin/env python3
# -*- coding:utf-8 -*-

# nuitka-project: --mingw64
# nuitka-project: --show-progress
# nuitka-project: --show-memory
# nuitka-project: --standalone
# nuitka-project: --windows-console-mode=disable
# nuitka-project: --windows-icon-from-ico=icon.ico
# nuitka-project: --output-dir=build

# 启用插件
# nuitka-project: --enable-plugin=tk-inter
# nuitka-project: --enable-plugin=upx

# 移除不必要的导入
# nuitka-project: --nofollow-import-to=numpy
# nuitka-project: --nofollow-import-to=pandas
"""
@author: Hsin-ming Chen
@license: MIT
@file: main.py
@time: 2025/08/16
@contact: hsinming.chen@gmail.com
@software: PyCharm
"""
import logging
from pathlib import Path
from time import sleep
import threading
from openpyxl import load_workbook
from helium import start_chrome, write, click, wait_until, find_all, kill_browser, Text, TextField, Button, RadioButton, CheckBox, Alert
from selenium.webdriver import ChromeOptions
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
import tkinter as tk
from tkinter import ttk, filedialog, StringVar, BooleanVar

# --- Global Constants ---
VERSION = "1.6"
EMEDICAL_URL = 'https://www.emedical.immi.gov.au/eMedUI/eMedical'
MAX_LOGIN_ATTEMPTS = 1

# --- Logging Setup ---
log_file = Path("log.txt")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler(log_file, mode='a', encoding='utf-8'),
    logging.StreamHandler()
])

# --- Global Event ---
stop_event = threading.Event()

# --- Class Definitions ---

class ExcelProcessor:
    def __init__(self):
        pass

    def extract_emedical_no(self, file_path):
        def is_black_font(cell):
            return (
                cell.font is None or
                cell.font.color is None or
                cell.font.color.rgb in ["FF000000", None]
            )

        def is_no_fill(cell):
            return (
                cell.fill is None or
                cell.fill.fgColor is None or
                cell.fill.fgColor.rgb in ["00000000", "FFFFFFFF", None]
            )

        if not Path(file_path).exists():
            logging.error(f"Excel file not found: {file_path}")
            return []

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            logging.error(f"Error reading Excel: {e}")
            return []

        all_emedical_nos = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found_header = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.strip() == "eMedical No.":
                        found_header = True
                        col_idx = cell.column
                        for r in range(cell.row + 1, ws.max_row + 1):
                            target_cell = ws.cell(row=r, column=col_idx)
                            if target_cell.value and is_black_font(target_cell) and is_no_fill(target_cell):
                                all_emedical_nos.append(str(target_cell.value))

            if not found_header:
                logging.warning(f"'eMedical No.' field not found in sheet {sheet_name}")

        wb.close()

        if len(all_emedical_nos) == 0:
            logging.warning("No valid eMedical No. read.")

        return all_emedical_nos

class EmedicalWebAutomator:
    def __init__(self, base_url):
        self.base_url = base_url
        self.options = ChromeOptions()
        self._setup_chrome_options()

    def _setup_chrome_options(self):
        self.options.add_argument("--disable-extensions")
        self.options.add_argument("--incognito")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument("--disable-background-networking")
        self.options.add_argument("--disable-component-update")
        self.options.add_argument("--disable-features=NetworkService,NetworkServiceInProcess")

    def login(self, user_id, password, headless=False):
        for attempt in range(1, MAX_LOGIN_ATTEMPTS + 1):
            try:
                logging.info("Starting browser and logging into eMedical system")
                start_chrome(self.base_url, headless=headless, options=self.options)
                write(user_id, into=TextField('User id'))
                write(password, into=TextField('Password'))
                click(Button('Logon'))
                wait_until(Text('Case search').exists, timeout_secs=10)
                logging.info("Login successful!")
                return True

            except NoSuchElementException:
                logging.error("Login fields not found, please check if the page has changed")
                break

            except TimeoutException:
                logging.warning(f"Login timed out, attempt {attempt}...")
                sleep(2)

            except WebDriverException as e:
                logging.error(f"Browser error: {e}")
                break

        logging.error("Login failed after multiple attempts, please check your credentials")
        return False

    def automate_cxr_exam(self, emed_no: str, country: str) -> bool:
        try:
            if not RadioButton('Using Health Case Identifier').is_selected():
                click(RadioButton('Using Health Case Identifier'))

            write(emed_no, into=TextField('ID'))
            click(Button('Search', to_right_of=Button('Reset')))
            wait_until(Text('Select:').exists)
            sleep(1)
            click(Button('All'))
            click(Button('Manage Case'))
            wait_until(Text('Pre exam: Health case details').exists)

            if Text('502 Chest X-Ray Examination').exists():
                click(Text('502 Chest X-Ray Examination'))

                if country == "美國":
                    click(Text('Findings'))
                    wait_until(Text('502 Chest X-Ray Examination: Findings').exists)
                    if not RadioButton('Normal', to_right_of=Text('Findings')).is_selected():
                        click(RadioButton('Normal', to_right_of=Text('Findings')))
                else:
                    click(Text('Detailed radiology findings'))
                    wait_until(Text('Detailed question').exists)
                    for normal_button in find_all(RadioButton('Normal')):
                        if not normal_button.is_selected():
                            click(normal_button)
                    if not RadioButton('Absent').is_selected():
                        click(RadioButton('Absent'))
                    if not RadioButton('No',
                                       to_right_of=RadioButton('Not selected', to_right_of=Text('7. Are there strong suspicions of active Tuberculosis (TB)?'))).is_selected():
                        click(RadioButton('No',
                                          to_right_of=RadioButton('Not selected', to_right_of=Text('7. Are there strong suspicions of active Tuberculosis (TB)?'))))
                    if country == "加拿大":
                        click(Button('Next'))
                        wait_until(Text('Special findings').exists)
                        if not RadioButton('None of the following are present').is_selected():
                            click(RadioButton('None of the following are present'))

                click(Button('Next'))
                wait_until(Text('502 Chest X-Ray Examination: Review exam details').exists)
                sleep(1)
                click(Button('Next'))

                if country == "美國":
                    wait_until(Text('502 Chest X-Ray Examination: Examiner Declaration').exists)
                    if Button('Prepare for declaration').exists() and Button('Prepare for declaration').is_enabled():
                        click(Button('Prepare for declaration'))
                else:
                    wait_until(Text('502 Chest X-Ray Examination: Grading & Examiner Declaration').exists)
                    if Button('Prepare for grading').exists() and Button('Prepare for grading').is_enabled():
                        click(Button('Prepare for grading'))

                wait_until(Text('Examiner declaration').exists)
                if not CheckBox(
                        'I declare that the chest X-ray examination report is a true and correct record of my findings.').is_checked():
                    click(CheckBox(
                        'I declare that the chest X-ray examination report is a true and correct record of my findings.'))

                if country != "美國":
                    if not RadioButton(
                            'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.').is_selected():
                        click(RadioButton(
                            'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.'))

                if Button('Submit Exam').exists() and Button('Submit Exam').is_enabled():
                    click(Button('Submit Exam'))
                    wait_until(Alert().exists)
                    Alert().accept()
                    wait_until(Text('Success').exists)

            click(Button('Close'))
            logging.info(f"Successfully processed ({country}): {emed_no}")
            return True

        except (NoSuchElementException, TimeoutException, WebDriverException) as e:
            logging.error(f"Automation failed for: {emed_no}, Error: {e}")
            try:
                if Button('Close').exists() and Button('Close').is_enabled():
                    click(Button('Close'))
            except Exception:
                pass
            return False

class EmedicalWorkflowManager:
    """
    A class to orchestrate the overall eMedical automation workflow.
    """
    def __init__(self, excel_processor, web_automator):
        self.excel_processor = excel_processor
        self.web_automator = web_automator
        # Callbacks for GUI updates
        self.update_status_callback = None
        self.update_emed_no_listbox_callback = None
        self.update_success_listbox_callback = None
        self.update_failure_listbox_callback = None
        self.update_counts_callback = None
        self.clear_listboxes_callback = None

    def set_gui_callbacks(self, update_status, update_emed_no_listbox, update_success_listbox, update_failure_listbox, update_counts, clear_listboxes):
        """Set the callback functions for GUI updates."""
        self.update_status_callback = update_status
        self.update_emed_no_listbox_callback = update_emed_no_listbox
        self.update_success_listbox_callback = update_success_listbox
        self.update_failure_listbox_callback = update_failure_listbox
        self.update_counts_callback = update_counts
        self.clear_listboxes_callback = clear_listboxes

    def start_workflow(self, user_id, password, excel_path, headless, close_browser):
        """Start the eMedical automation workflow."""
        if not self.update_status_callback:
            logging.error("GUI update callbacks are not set.")
            return

        self.update_status_callback(f"Reading eMedical No. from {excel_path}")
        if not Path(excel_path).exists():
            self.update_status_callback(f"Error: File {excel_path} not found")
            return

        emedical_numbers = self.excel_processor.extract_emedical_no(Path(excel_path))
        if not emedical_numbers:
            self.update_status_callback("No eMedical No. read.")
            return
        self.update_status_callback(f"Read {len(emedical_numbers)} eMedical No.")

        # Clear and populate the eMedical No. listbox
        if self.clear_listboxes_callback:
            self.clear_listboxes_callback()
        for emed_no in emedical_numbers:
            if self.update_emed_no_listbox_callback:
                self.update_emed_no_listbox_callback(emed_no)

        if not self.web_automator.login(user_id, password, headless):
            self.update_status_callback("Login failed, please check your credentials")
            return

        for index, emed_no in enumerate(emedical_numbers):
            if stop_event.is_set():
                self.update_status_callback("Processing stopped by user")
                break

            self.update_status_callback(f'Processing: {emed_no}')
            logging.info(f'Processing: {emed_no}')

            if self.update_emed_no_listbox_callback:
                self.update_emed_no_listbox_callback(emed_no, index=index, highlight=True)

            country = self._get_country(emed_no)
            success = False

            if country == "未知國家":
                logging.warning(f"Unknown country for eMedical No.: {emed_no}")
            else:
                success = self.web_automator.automate_cxr_exam(emed_no, country)

            if self.update_emed_no_listbox_callback:
                self.update_emed_no_listbox_callback(emed_no, index=index, highlight=False)

            if success:
                if self.update_success_listbox_callback:
                    self.update_success_listbox_callback(emed_no)
            else:
                if self.update_failure_listbox_callback:
                    self.update_failure_listbox_callback(emed_no)

            if self.update_counts_callback:
                self.update_counts_callback()

        if self.update_counts_callback:
            self.update_counts_callback() # Ensure final counts are updated on GUI

        self.update_status_callback("Processing complete!")
        logging.info("Processing complete!")

        if close_browser or headless:
            logging.info("Closing browser")
            kill_browser()

    def stop_workflow(self):
        """Trigger the stop event to halt the processing."""
        stop_event.set()

    def _get_country(self, emed_no):
        """Determine the country based on the eMedical No. prefix."""
        prefix_to_country = {
            "HAP": "澳大利亞",
            "TRN": "澳大利亞",
            "NZER": "紐西蘭",
            "NZHR": "紐西蘭",
            "IME": "加拿大",
            "UMI": "加拿大",
            "UCI": "加拿大",
            "CEAC": "美國",
        }

        for prefix, country in prefix_to_country.items():
            if emed_no.startswith(prefix):
                return country

        return "未知國家"

class EmedicalGUI:
    """
    A class to manage the Tkinter GUI for the eMedical automation tool.
    """
    def __init__(self, master, workflow_manager):
        self.master = master
        self.workflow_manager = workflow_manager

        # Tkinter variables
        self.user_id_var = StringVar()
        self.password_var = StringVar()
        self.excel_path_var = StringVar()
        self.headless_var = BooleanVar()
        self.close_browser_var = BooleanVar()
        self.status_var = StringVar()

        # Listbox and Label references
        self.emed_no_listbox = None
        self.success_listbox = None
        self.failure_listbox = None
        self.success_label = None
        self.failure_label = None

        # Pass GUI update methods as callbacks to the workflow manager
        self.workflow_manager.set_gui_callbacks(
            update_status=self.update_status,
            update_emed_no_listbox=self.update_emed_no_listbox,
            update_success_listbox=self.update_success_listbox,
            update_failure_listbox=self.update_failure_listbox,
            update_counts=self.update_counts,
            clear_listboxes=self.clear_listboxes
        )

    def setup_ui(self):
        """Build the entire GUI interface."""
        self.master.title(f"eMedical Automation v{VERSION} By HsinMing Chen")
        self.master.geometry("500x700")
        self.master.minsize(500, 700)
        self.master.resizable(True, True)

        style = ttk.Style()
        style.configure("TButton", font=("Arial", 12))
        style.configure("TLabel", font=("Arial", 11))
        style.configure("TEntry", font=("Arial", 11))

        main_frame = ttk.Frame(self.master, padding=10)
        main_frame.grid(row=0, column=0, sticky="nsew")

        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)

        for i in range(8):
            main_frame.rowconfigure(i, weight=1)
        main_frame.columnconfigure(0, weight=1)

        # User Login Frame
        user_frame = ttk.LabelFrame(main_frame, text="User Login", padding=10)
        user_frame.grid(row=0, column=0, sticky="nsew", pady=5)
        ttk.Label(user_frame, text="User ID:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(user_frame, textvariable=self.user_id_var, width=30).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(user_frame, text="Password:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(user_frame, textvariable=self.password_var, show='*', width=30).grid(row=1, column=1, padx=5, pady=5)

        # Excel File Frame
        file_frame = ttk.LabelFrame(main_frame, text="Excel File", padding=10)
        file_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        file_entry = ttk.Entry(file_frame, textvariable=self.excel_path_var)
        file_entry.grid(row=0, column=0, sticky="ew", padx=5)
        ttk.Button(file_frame, text="Browse", command=self._select_file).grid(row=0, column=1, padx=5, sticky="e")
        file_frame.columnconfigure(0, weight=1)

        # Options Frame
        options_frame = ttk.LabelFrame(main_frame, text="Options", padding=10)
        options_frame.grid(row=2, column=0, sticky="nsew", pady=5)
        ttk.Checkbutton(options_frame, text="Headless Mode", variable=self.headless_var).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(options_frame, text="Kill Browser After Completion", variable=self.close_browser_var).grid(row=1, column=0, sticky="w")

        # Status Label
        ttk.Label(main_frame, textvariable=self.status_var, foreground='blue').grid(row=3, column=0, pady=5, sticky="w")

        # eMedical No. Listbox Frame
        list_frame = ttk.LabelFrame(main_frame, text="eMedical No. 清單", padding=10)
        list_frame.grid(row=4, column=0, sticky="nsew", pady=5)
        self.emed_no_listbox = tk.Listbox(list_frame, height=5)
        emed_no_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.emed_no_listbox.yview)
        self.emed_no_listbox.config(yscrollcommand=emed_no_scrollbar.set)
        self.emed_no_listbox.grid(row=0, column=0, sticky="nsew")
        emed_no_scrollbar.grid(row=0, column=1, sticky="ns")
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        # Results Frame
        result_frame = ttk.LabelFrame(main_frame, text="Processing Results", padding=10)
        result_frame.grid(row=5, column=0, sticky="nsew", pady=5)
        self.success_label = ttk.Label(result_frame, text="成功的 eMedical No. (數量: 0)")
        self.success_label.grid(row=0, column=0, sticky="w")
        self.success_listbox = tk.Listbox(result_frame, height=5)
        self.success_listbox.grid(row=1, column=0, sticky="nsew", padx=5, pady=2)
        self.failure_label = ttk.Label(result_frame, text="失敗的 eMedical No. (數量: 0)")
        self.failure_label.grid(row=2, column=0, sticky="w")
        self.failure_listbox = tk.Listbox(result_frame, height=5)
        self.failure_listbox.grid(row=3, column=0, sticky="nsew", padx=5, pady=2)
        result_frame.rowconfigure(1, weight=1)
        result_frame.rowconfigure(3, weight=1)
        result_frame.columnconfigure(0, weight=1)

        # Button Frame
        button_frame = ttk.Frame(main_frame, padding=10)
        button_frame.grid(row=6, column=0, sticky="nsew", pady=5)
        ttk.Button(button_frame, text="Start", command=self._start_automation_thread).grid(row=0, column=0, sticky="ew")
        ttk.Button(button_frame, text="Stop", command=self._stop_automation).grid(row=0, column=1, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

    def _select_file(self):
        """Open a file dialog to select an Excel file."""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.excel_path_var.set(file_path)

    def _start_automation_thread(self):
        """Start the automation workflow in a separate thread."""
        user_id = self.user_id_var.get()
        password = self.password_var.get()
        excel_path = self.excel_path_var.get()
        headless = self.headless_var.get()
        close_browser = self.close_browser_var.get()

        if not user_id or not password or not excel_path:
            self.update_status("Please fill in all fields!")
            return

        self.clear_listboxes()
        self.update_counts()
        self.update_status("Processing...")

        threading.Thread(
            target=self.workflow_manager.start_workflow,
            args=(user_id, password, excel_path, headless, close_browser),
            daemon=True
        ).start()

    def _stop_automation(self):
        """Stop the current automation process."""
        self.workflow_manager.stop_workflow()
        self.update_status("Processing stopped")

    # --- GUI Update Callbacks ---
    def update_status(self, msg):
        """Update the status label."""
        self.status_var.set(msg)
        self.master.after(100, lambda: self.master.update_idletasks())

    def update_emed_no_listbox(self, emed_no, index=None, highlight=False):
        """Update the eMedical No. listbox."""
        if index is None:
            self.emed_no_listbox.insert(tk.END, emed_no)
        else:
            self.emed_no_listbox.itemconfig(index, {'bg': 'blue' if highlight else 'white', 'fg': 'white' if highlight else 'black'})

    def update_success_listbox(self, emed_no):
        """Add an item to the success listbox."""
        self.success_listbox.insert(tk.END, emed_no)

    def update_failure_listbox(self, emed_no):
        """Add an item to the failure listbox."""
        self.failure_listbox.insert(tk.END, emed_no)

    def update_counts(self):
        """Update the success/failure count labels."""
        success_count = self.success_listbox.size()
        failure_count = self.failure_listbox.size()
        self.success_label.config(text=f"成功的 eMedical No. (數量: {success_count})")
        self.failure_label.config(text=f"失敗的 eMedical No. (數量: {failure_count})")

    def clear_listboxes(self):
        """Clear all listboxes."""
        self.emed_no_listbox.delete(0, tk.END)
        self.success_listbox.delete(0, tk.END)
        self.failure_listbox.delete(0, tk.END)

# --- Main Execution ---
if __name__ == "__main__":
    root = tk.Tk()
    excel_processor = ExcelProcessor()
    web_automator = EmedicalWebAutomator(EMEDICAL_URL)
    workflow_manager = EmedicalWorkflowManager(excel_processor, web_automator)
    app_gui = EmedicalGUI(root, workflow_manager)
    app_gui.setup_ui()
    root.mainloop()

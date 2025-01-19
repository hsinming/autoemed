#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
@author: Hsin-ming Chen
@license: GPL
@file: main-command-line.py
@time: 2025/01/19
@contact: hsinming.chen@gmail.com
@software: PyCharm
"""
import logging
import argparse
from pathlib import Path
from time import sleep
import openpyxl
from helium import start_chrome, write, click, wait_until, find_all, kill_browser, Text, TextField, Button, RadioButton, CheckBox, Alert


EMEDICAL_URL = 'https://www.emedical.immi.gov.au/eMedUI/eMedical'


# 設定日誌
log_file = Path("log.txt")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler(log_file, mode='a', encoding='utf-8'),
    logging.StreamHandler()
])


def parse_args():
    parser = argparse.ArgumentParser(description="eMedical 自動化處理程式")
    parser.add_argument("--user_id", required=True, help="eMedical 登入帳號")
    parser.add_argument("--password", required=True, help="eMedical 密碼")
    parser.add_argument("--excel_path", required=True, type=Path, help="Excel 檔案路徑")
    parser.add_argument("--headless", type=bool, default=False, help="是否使用 headless 模式")
    parser.add_argument("--kill_browser", type=bool, default=True, help="執行結束後是否關閉瀏覽器")
    return parser.parse_args()


def is_black_font(cell):
    return cell.font is None or cell.font.color is None or cell.font.color.rgb in ["FF000000", None]


def is_no_fill(cell):
    return cell.fill is None or cell.fill.fgColor is None or cell.fill.fgColor.rgb in ["00000000", "FFFFFFFF", None]


def extract_emedical_no(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_emedical_nos = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip() == "eMedical No.":
                    # 獲取此列下方的所有值
                    col_idx = cell.column
                    for r in range(cell.row + 1, ws.max_row + 1):
                        target_cell = ws.cell(row=r, column=col_idx)
                        if target_cell.value and is_black_font(target_cell) and is_no_fill(target_cell):
                            all_emedical_nos.append(target_cell.value)

    wb.close()
    return all_emedical_nos


def process_case(emed_no: str, country: str):
    """
    通用的 eMedical 案例處理函數
    根據 country 參數決定是否有額外的步驟
    """
    try:
        click(RadioButton('Using Health Case Identifier'))
        write(emed_no, into=TextField('ID'))
        click(Button('Search'))
        wait_until(Text('Select:').exists)
        sleep(1)
        click(Button('All'))
        click(Button('Manage Case'))

        wait_until(Text('Pre exam: Health case details').exists)
        if Text('502 Chest X-Ray Examination').exists():
            click(Text('502 Chest X-Ray Examination'))

            # 依國家需求點擊不同的按鈕
            if country == "美國":
                click(Text('Findings'))

                wait_until(Text('502 Chest X-Ray Examination: Findings').exists)
                if not RadioButton('Normal', to_right_of=Text('Findings')).is_selected():
                    click(RadioButton('Normal', to_right_of=Text('Findings')))

            else:
                click(Text('Detailed radiology findings'))

                wait_until(Text('Detailed question').exists)
                for rb_normal in find_all(RadioButton('Normal')):
                    if not rb_normal.is_selected():
                        click(rb_normal)

                if not RadioButton('Absent').is_selected():
                    click(RadioButton('Absent'))

                if not RadioButton('No').is_selected():
                    click(RadioButton('No'))

                if country == "加拿大":
                    click(Button('Next'))
                    wait_until(Text('Special findings').exists)
                    if not RadioButton('None of the following are present').is_selected():
                        click(RadioButton('None of the following are present'))

            click(Button('Next'))
            wait_until(Text('502 Chest X-Ray Examination: Review exam details').exists)
            sleep(1)
            click(Button('Next'))

            if country == "美國":
                wait_until(Text('502 Chest X-Ray Examination: Examiner Declaration').exists)
                if Button('Prepare for declaration').exists() and Button('Prepare for declaration').is_enabled():
                    click(Button('Prepare for declaration'))
            else:
                wait_until(Text('502 Chest X-Ray Examination: Grading & Examiner Declaration').exists)
                if Button('Prepare for grading').exists() and Button('Prepare for grading').is_enabled():
                    click(Button('Prepare for grading'))

            wait_until(Text('Examiner declaration').exists)
            if not CheckBox(
                    'I declare that the chest X-ray examination report is a true and correct record of my findings.').is_checked():
                click(CheckBox(
                    'I declare that the chest X-ray examination report is a true and correct record of my findings.'))

            if country != "美國":
                if not RadioButton(
                        'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.').is_selected():
                    click(RadioButton(
                        'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.'))

            if Button('Submit Exam').exists() and Button('Submit Exam').is_enabled():
                click(Button('Submit Exam'))
                wait_until(Alert().exists)
                Alert().accept()
                wait_until(Text('Success').exists)

        click(Button('Close'))
        logging.info(f"成功處理 ({country}): {emed_no}")
        return True

    except Exception as e:
        logging.error(f"處理失敗 ({country}): {emed_no}, 錯誤: {e}")
        return False


if __name__ == "__main__":
    args = parse_args()
    logging.info(f"讀取 {args.excel_path}")
    emedical_numbers = extract_emedical_no(args.excel_path)
    logging.info(f"讀取 {len(emedical_numbers)} 個 eMedical No.")

    country_map = {
        'HAP': "澳大利亞",
        'TRN': "澳大利亞",
        'NZER': "紐西蘭",
        'NZHR': "紐西蘭",
        'IME': "加拿大",
        'UMI': "加拿大",
        'UCI': "加拿大",
        'CEAC': "美國"
    }

    success_list = []
    failure_list = []

    start_chrome(EMEDICAL_URL, headless=args.headless)
    write(args.user_id, into=TextField('User id'))
    write(args.password, into=TextField('Password'))
    click(Button('Logon'))

    wait_until(Text('Case search').exists, timeout_secs=10)

    for emed_no in emedical_numbers:
        logging.info(f'現在處理: {emed_no}')

        country = next((v for k, v in country_map.items() if emed_no.startswith(k)), None)

        if country:
            success = process_case(emed_no, country)
        else:
            logging.warning(f"未知的 eMedical No. 類型: {emed_no}")
            success = False

        if success:
            success_list.append(emed_no)
        else:
            failure_list.append(emed_no)

    logging.info("處理完成！")
    logging.info(f"成功: {len(success_list)}, 失敗: {len(failure_list)}")

    if args.kill_browser:
        kill_browser()

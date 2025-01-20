#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
@author: Hsin-ming Chen
@license: GPL
@file: main-gui.py
@time: 2025/01/20
@contact: hsinming.chen@gmail.com
@software: PyCharm
"""
import logging
from pathlib import Path
from time import sleep
import threading
from openpyxl import load_workbook
from helium import start_chrome, write, click, wait_until, find_all, kill_browser, Text, TextField, Button, RadioButton, CheckBox, Alert
import tkinter as tk
from tkinter import ttk, filedialog, StringVar, BooleanVar

EMEDICAL_URL = 'https://www.emedical.immi.gov.au/eMedUI/eMedical'

# 設定日誌
log_file = Path("log.txt")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler(log_file, mode='a', encoding='utf-8'),
    logging.StreamHandler()
])

# 中止標誌 (使用 threading.Event)
stop_event = threading.Event()


def extract_emedical_no(file_path):
    """從 Excel 讀取 eMedical No. 清單"""
    def is_black_font(cell):
        return cell.font is None or cell.font.color is None or cell.font.color.rgb in ["FF000000", None]

    def is_no_fill(cell):
        return cell.fill is None or cell.fill.fgColor is None or cell.fill.fgColor.rgb in ["00000000", "FFFFFFFF", None]

    wb = load_workbook(file_path, data_only=True)
    all_emedical_nos = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip() == "eMedical No.":
                    # 獲取此列下方的所有值
                    col_idx = cell.column
                    for r in range(cell.row + 1, ws.max_row + 1):
                        target_cell = ws.cell(row=r, column=col_idx)
                        if target_cell.value and is_black_font(target_cell) and is_no_fill(target_cell):
                            all_emedical_nos.append(target_cell.value)

    wb.close()
    return all_emedical_nos


def emedical_cxr_automation(emed_no: str, country: str):
    """
    通用的 eMedical 案例處理函數
    根據 country 參數決定是否有額外的步驟
    """
    try:
        if not RadioButton('Using Health Case Identifier').is_selected():
            click(RadioButton('Using Health Case Identifier'))
        write(emed_no, into=TextField('ID'))
        click(Button('Search'))

        wait_until(Text('Select:').exists)
        sleep(1)    #wait for reading status
        click(Button('All'))
        click(Button('Manage Case'))

        wait_until(Text('Pre exam: Health case details').exists)
        if Text('502 Chest X-Ray Examination').exists():
            click(Text('502 Chest X-Ray Examination'))

            # 依國家需求點擊不同的按鈕
            if country == "美國":
                click(Text('Findings'))

                wait_until(Text('502 Chest X-Ray Examination: Findings').exists)
                if not RadioButton('Normal', to_right_of=Text('Findings')).is_selected():
                    click(RadioButton('Normal', to_right_of=Text('Findings')))

            else:
                click(Text('Detailed radiology findings'))

                wait_until(Text('Detailed question').exists)
                for rb_normal in find_all(RadioButton('Normal')):
                    if not rb_normal.is_selected():
                        click(rb_normal)

                if not RadioButton('Absent').is_selected():
                    click(RadioButton('Absent'))

                if not RadioButton('No').is_selected():
                    click(RadioButton('No'))

                if country == "加拿大":
                    click(Button('Next'))
                    wait_until(Text('Special findings').exists)
                    if not RadioButton('None of the following are present').is_selected():
                        click(RadioButton('None of the following are present'))

            click(Button('Next'))
            wait_until(Text('502 Chest X-Ray Examination: Review exam details').exists)
            sleep(1)
            click(Button('Next'))

            if country == "美國":
                wait_until(Text('502 Chest X-Ray Examination: Examiner Declaration').exists)
                if Button('Prepare for declaration').exists() and Button('Prepare for declaration').is_enabled():
                    click(Button('Prepare for declaration'))
            else:
                wait_until(Text('502 Chest X-Ray Examination: Grading & Examiner Declaration').exists)
                if Button('Prepare for grading').exists() and Button('Prepare for grading').is_enabled():
                    click(Button('Prepare for grading'))

            wait_until(Text('Examiner declaration').exists)
            if not CheckBox(
                    'I declare that the chest X-ray examination report is a true and correct record of my findings.').is_checked():
                click(CheckBox(
                    'I declare that the chest X-ray examination report is a true and correct record of my findings.'))

            if country != "美國":
                if not RadioButton(
                        'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.').is_selected():
                    click(RadioButton(
                        'A - No evidence of active TB, or changes consistent with old or inactive TB, or changes suggestive of other significant diseases identified.'))

            if Button('Submit Exam').exists() and Button('Submit Exam').is_enabled():
                click(Button('Submit Exam'))
                wait_until(Alert().exists)
                Alert().accept()
                wait_until(Text('Success').exists)

        click(Button('Close'))
        logging.info(f"成功處理 ({country}): {emed_no}")
        return True

    except Exception as e:
        logging.error(f"處理失敗 ({country}): {emed_no}, 錯誤: {e}")
        return False


def emedical_workflow(root, user_id, password, excel_path, status_var, emed_no_listbox, success_listbox, failure_listbox, headless, close_browser, update_counts):
    logging.info(f"讀取 eMedical No. 來自 {excel_path}")
    emedical_numbers = extract_emedical_no(Path(excel_path))
    logging.info(f"讀取 {len(emedical_numbers)} 個 eMedical No.")

    # 將 eMedical No. 加入 GUI 清單
    for emed_no in emedical_numbers:
        emed_no_listbox.insert(tk.END, emed_no)

    country_map = {
        'HAP': "澳大利亞",
        'TRN': "澳大利亞",
        'NZER': "紐西蘭",
        'NZHR': "紐西蘭",
        'IME': "加拿大",
        'UMI': "加拿大",
        'UCI': "加拿大",
        'CEAC': "美國"
    }

    try:
        start_chrome(EMEDICAL_URL, headless=headless)
        write(user_id, into=TextField('User id'))
        write(password, into=TextField('Password'))
        click(Button('Logon'))
        wait_until(Text('Case search').exists, timeout_secs=10)

        for index, emed_no in enumerate(emedical_numbers):
            if stop_event.is_set():
                logging.info("用戶手動中止處理")
                status_var.set("處理已中止")
                break

            status_var.set(f'現在處理: {emed_no}')
            logging.info(f'現在處理: {emed_no}')
            root.update()

            # 標記當前處理中的 eMedical No.
            emed_no_listbox.itemconfig(index, {'bg': 'blue', 'fg': 'white'})

            country = next((v for k, v in country_map.items() if emed_no.startswith(k)), None)

            if country:
                success = emedical_cxr_automation(emed_no, country)
            else:
                logging.warning(f"未知國家的 eMedical No.: {emed_no}")
                success = False

            # 恢復 eMedical No. 顏色
            emed_no_listbox.itemconfig(index, {'bg': 'white', 'fg': 'black'})

            # 根據結果分類
            if success:
                success_listbox.insert(tk.END, emed_no)
            else:
                failure_listbox.insert(tk.END, emed_no)

            update_counts()

        status_var.set(f"處理完成！成功: {success_listbox.size()}, 失敗: {failure_listbox.size()}")
        logging.info(f"處理完成！成功: {success_listbox.size()}, 失敗: {failure_listbox.size()}")

    except Exception as e:
        logging.error(f"eMedical 工作流程出錯: {e}")
        status_var.set("處理時發生錯誤")

    finally:
        if close_browser:
            kill_browser()
        logging.info("處理流程結束")


def start_gui():
    root = tk.Tk()
    root.title("eMedical Automation v1.1 By HsinMing Chen")
    root.geometry("500x800")  # 設定視窗大小
    root.resizable(False, False)  # 禁止調整大小
    root.attributes('-topmost', True)  # 永遠在最上層

    style = ttk.Style()
    style.configure("TButton", font=("Arial", 12))
    style.configure("TLabel", font=("Arial", 11))
    style.configure("TEntry", font=("Arial", 11))

    main_frame = ttk.Frame(root, padding=10)
    main_frame.pack(fill=tk.BOTH, expand=True)

    # 使用者資訊區塊
    user_frame = ttk.LabelFrame(main_frame, text="User Login", padding=10)
    user_frame.pack(fill=tk.X, pady=5)

    ttk.Label(user_frame, text="User ID:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    user_id_var = StringVar()
    ttk.Entry(user_frame, textvariable=user_id_var, width=30).grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(user_frame, text="Password:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    password_var = StringVar()
    ttk.Entry(user_frame, textvariable=password_var, show='*', width=30).grid(row=1, column=1, padx=5, pady=5)

    # Excel 檔案選擇區塊
    file_frame = ttk.LabelFrame(main_frame, text="Excel File", padding=10)
    file_frame.pack(fill=tk.X, pady=5)

    excel_path_var = StringVar()
    ttk.Entry(file_frame, textvariable=excel_path_var, width=40).pack(side=tk.LEFT, padx=5)

    def select_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        excel_path_var.set(file_path)

    ttk.Button(file_frame, text="Browse", command=select_file).pack(side=tk.RIGHT, padx=5)

    # 選項區塊
    options_frame = ttk.LabelFrame(main_frame, text="Options", padding=10)
    options_frame.pack(fill=tk.X, pady=5)

    headless_var = BooleanVar()
    close_browser_var = BooleanVar()
    ttk.Checkbutton(options_frame, text="Headless Mode", variable=headless_var).pack(anchor="w")
    ttk.Checkbutton(options_frame, text="Kill Browser After Completion", variable=close_browser_var).pack(anchor="w")

    # 狀態顯示
    status_var = StringVar()
    ttk.Label(main_frame, textvariable=status_var, foreground='blue').pack(pady=5)

    # eMedical No. Listbox
    list_frame = ttk.LabelFrame(main_frame, text="eMedical No. 清單", padding=10)
    list_frame.pack(fill=tk.BOTH, expand=True, pady=5)

    emed_no_listbox = tk.Listbox(list_frame, height=5, width=50)
    emed_no_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=emed_no_listbox.yview)
    emed_no_listbox.config(yscrollcommand=emed_no_scrollbar.set)
    emed_no_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    emed_no_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # 成功 & 失敗 ListBox
    result_frame = ttk.LabelFrame(main_frame, text="Processing Results", padding=10)
    result_frame.pack(fill=tk.BOTH, expand=True, pady=5)

    ttk.Label(result_frame, text="成功的 eMedical No. (數量: 0)", name="success_label").pack(anchor="w")
    success_listbox = tk.Listbox(result_frame, height=5, width=25)
    success_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

    ttk.Label(result_frame, text="失敗的 eMedical No. (數量: 0)", name="failure_label").pack(anchor="w")
    failure_listbox = tk.Listbox(result_frame, height=5, width=25)
    failure_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

    def update_counts():
        success_count = success_listbox.size()
        failure_count = failure_listbox.size()

        success_label = result_frame.nametowidget("success_label")
        failure_label = result_frame.nametowidget("failure_label")
        success_label.config(text=f"成功的 eMedical No. (數量: {success_count})")
        failure_label.config(text=f"失敗的 eMedical No. (數量: {failure_count})")

    def start_emedical_workflow():
        user_id = user_id_var.get()
        password = password_var.get()
        excel_path = excel_path_var.get()
        headless = headless_var.get()
        close_browser = close_browser_var.get()

        if not user_id or not password or not excel_path:
            status_var.set("請填寫所有欄位！")
            return

        status_var.set("正在處理...")
        root.update()

        stop_event.clear()

        # **使用 Thread 避免 GUI 被阻塞**
        emedical_workflow_thread = threading.Thread(
            target=emedical_workflow,
            args=(root, user_id, password, excel_path, status_var, emed_no_listbox, success_listbox, failure_listbox, headless, close_browser, update_counts),
            daemon=True  # 設為 Daemon，確保主程式關閉時該 Thread 也會終止
        )
        emedical_workflow_thread.start()

    def stop_emedical_workflow():
        """中止處理流程"""
        stop_event.set()
        status_var.set("已中止處理")

    # 按鈕區塊
    button_frame = ttk.Frame(main_frame, padding=10)
    button_frame.pack(fill=tk.X, pady=5)

    ttk.Button(button_frame, text="Start", command=start_emedical_workflow, style="TButton").pack(side=tk.LEFT,
                                                                                                  expand=True)
    ttk.Button(button_frame, text="Stop", command=stop_emedical_workflow, style="TButton").pack(side=tk.RIGHT,
                                                                                                expand=True)

    root.mainloop()


if __name__ == "__main__":
    start_gui()
